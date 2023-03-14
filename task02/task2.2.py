import openpyxl


# Mở file excel và lấy sheet cần thao tác
workbook = openpyxl.load_workbook('G:\input.xlsx')
sheet = workbook['MAU']

# Lấy dữ liệu và sắp xếp theo thứ tự tên
data = []
for row in range(2, sheet.max_row + 1):
    name = sheet.cell(row=row, column=2).value
    math = sheet.cell(row=row, column=3).value
    physics = sheet.cell(row=row, column=4).value
    chemistry = sheet.cell(row=row, column=5).value
    data.append({'name': name, 'math': math, 'physics': physics, 'chemistry': chemistry})

data.sort(key=lambda x: x['name'])

# Tính điểm trung bình và phân loại học lực
for row, student in enumerate(data, start=2):
    math = student['math']
    physics = student['physics']
    chemistry = student['chemistry']
    average = (math + physics + chemistry) / 3
    sheet.cell(row=row, column=6).value = average

    if average >= 8:
        sheet.cell(row=row, column=7).value = 'Giỏi'
    elif average >= 6.5:
        sheet.cell(row=row, column=7).value = 'Khá'
    else:
        sheet.cell(row=row, column=7).value = 'Trung bình'

# Lưu file excel
workbook.save('output.xlsx')